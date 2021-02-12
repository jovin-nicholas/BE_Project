# -*- coding: utf-8 -*-
"""
Created on Tue Feb  9 20:37:47 2021

@author: akshay
"""


import openpyxl 
import math

# Give the location of the file 
path = "SoftSkillData.xlsx"

# workbook object is created 
wb_obj = openpyxl.load_workbook(path) 
sheet_obj = wb_obj.active
max_rows = sheet_obj.max_row

for i in range(2,max_rows):
    cptr = sheet_obj.cell(row = i, column = 4).value
    c1 = sheet_obj.cell(row = i, column = 6).value
    c2 = sheet_obj.cell(row = i, column = 7).value
    c3 = sheet_obj.cell(row = i, column = 8).value
    c4 = sheet_obj.cell(row = i, column = 9).value
    c5 = sheet_obj.cell(row = i, column = 10).value
    c6 = sheet_obj.cell(row = i, column = 11).value
    c7 = sheet_obj.cell(row = i, column = 12).value
    c8 = sheet_obj.cell(row = i, column = 13).value
    c9 = sheet_obj.cell(row = i, column = 14).value
    c10 = sheet_obj.cell(row = i, column = 15).value
    
    
    o1 = sheet_obj.cell(row = i, column = 21)
    o2 = sheet_obj.cell(row = i, column = 22)
    o3 = sheet_obj.cell(row = i, column = 23)
    o4 = sheet_obj.cell(row = i, column = 24)
    
    o1.value = math.ceil((c5*20)/25)
    o2.value = math.ceil(cptr*20/10)
    o3.value = math.ceil(c3*20/25)
    total = c1+c2+c4+c6+c7+c8+c9+c10
    o4.value = math.ceil(total*20/200)
    
wb_obj.save("SoftSkillData.xlsx") 